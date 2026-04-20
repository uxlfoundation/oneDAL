#!/usr/bin/env python3
"""Generate Excel benchmark report for HDBSCAN sprint review."""
import os
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = "scripts/benchmark_results"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Full results from extended benchmark run (2026-04-20)
# (name, n, d, mcs, ms, sk_ms, sk_c, cpu_ms, cpu_c, cpu_ari, cpu_spd, gpu_ms, gpu_c, gpu_ari, gpu_spd, notes)
results = [
    ("blobs_500_2d", 500, 2, 5, 3, 6.7, 3, 3.1, 3, 1.0, 2.18, 4.9, 3, 1.0, 1.38, ""),
    ("blobs_1k_2d", 1000, 2, 10, 5, 18.4, 4, 7.0, 4, 0.997, 2.63, 10.4, 4, 1.0, 1.77, ""),
    ("moons_2k_2d", 2000, 2, 15, 10, 112.8, 2, 21.0, 2, 1.0, 5.37, 30.0, 2, 1.0, 3.76, "non-convex shape"),
    ("blobs_5k_2d", 5000, 2, 15, 10, 795.7, 5, 130.9, 5, 1.0, 6.08, 161.6, 5, 1.0, 4.92, ""),
    ("blobs_5k_5d", 5000, 5, 15, 10, 788.4, 6, 131.3, 6, 0.999, 6.0, 142.4, 6, 1.0, 5.54, ""),
    ("blobs_5k_10d", 5000, 10, 15, 10, 899.9, 7, 133.1, 7, 1.0, 6.76, 133.0, 7, 1.0, 6.77, "best GPU case"),
    ("blobs_10k_2d", 10000, 2, 20, 15, 2818.7, 7, 270.1, 7, 0.999, 10.44, 669.2, 7, 1.0, 4.21, "best CPU case"),
    ("blobs_10k_10d", 10000, 10, 20, 15, 1890.2, 8, 277.5, 8, 1.0, 6.81, 692.5, 8, 1.0, 2.73, ""),
    ("blobs_20k_2d", 20000, 2, 25, 15, 8272.4, 7, 1117.2, 7, 1.0, 7.40, 1981.3, 7, 0.996, 4.18, ""),
    ("blobs_20k_5d", 20000, 5, 25, 15, 6640.6, 9, 1120.1, 9, 1.0, 5.93, None, None, None, None, "GPU: float32 precision loss at scale"),
    ("blobs_30k_2d", 30000, 2, 30, 20, 18448.0, 7, 2656.2, 7, 1.0, 6.95, None, None, None, None, "GPU: float32 precision loss at scale"),
    ("blobs_50k_2d", 50000, 2, 40, 25, 55114.0, 6, 7288.0, 6, 1.0, 7.56, None, None, None, None, "GPU: OOM (10GB dist matrix)"),
    ("hdbscan_dense_10k", 10500, 2, 15, 10, 1808.7, 7, 306.1, 7, 1.0, 5.91, 792.6, 7, 1.0, 2.28, "reference dataset"),
]

wb = Workbook()

# --- Sheet 1: Full Results ---
ws = wb.active
ws.title = "Benchmark Results"

ws["A1"] = "HDBSCAN Performance Benchmark: oneDAL vs scikit-learn"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "Date: 2026-04-20 | CPU: Xeon 8580 (56 cores) | GPU: Data Center GPU Max 1550 | OS: Linux"
ws["A2"].font = Font(size=10, italic=True)
ws["A3"] = "oneDAL CPU uses DAAL kernel (BLAS GEMM + threader_for + ISA dispatch). GPU uses SYCL kernels."
ws["A3"].font = Font(size=10, italic=True)

headers = [
    "Dataset", "N", "D", "min_cluster_size", "min_samples",
    "sklearn (ms)", "sklearn clusters",
    "oneDAL CPU (ms)", "CPU clusters", "CPU ARI", "CPU Speedup vs sklearn",
    "oneDAL GPU (ms)", "GPU clusters", "GPU ARI", "GPU Speedup vs sklearn",
    "Notes"
]
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=10)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=5, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border

for row_idx, row in enumerate(results, 6):
    name, n, d, mcs, ms, sk_ms, sk_c, cpu_ms, cpu_c, cpu_ari, cpu_spd, gpu_ms, gpu_c, gpu_ari, gpu_spd, notes = row
    values = [
        name, n, d, mcs, ms,
        round(sk_ms, 1), sk_c,
        round(cpu_ms, 1) if cpu_ms else "N/A",
        cpu_c if cpu_c else "N/A",
        round(cpu_ari, 4) if cpu_ari is not None else "N/A",
        round(cpu_spd, 2) if cpu_spd else "N/A",
        round(gpu_ms, 1) if gpu_ms else "N/A",
        gpu_c if gpu_c else "N/A",
        round(gpu_ari, 4) if gpu_ari is not None else "N/A",
        round(gpu_spd, 2) if gpu_spd else "N/A",
        notes,
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        # Highlight speedup cells
        if col_idx == 11 and isinstance(val, float):
            if val >= 8.0:
                cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            elif val >= 6.0:
                cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            elif val >= 4.0:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        if col_idx == 15 and isinstance(val, float):
            if val >= 6.0:
                cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            elif val >= 5.0:
                cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            elif val >= 3.0:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# Summary row
summary_row = 6 + len(results) + 1
ws.cell(row=summary_row, column=1, value="AVERAGE").font = Font(bold=True, size=11)
cpu_spds = [r[10] for r in results if r[10]]
gpu_spds = [r[14] for r in results if r[14]]
ws.cell(row=summary_row, column=11, value=round(np.mean(cpu_spds), 2)).font = Font(bold=True, size=12, color="006100")
if gpu_spds:
    ws.cell(row=summary_row, column=15, value=round(np.mean(gpu_spds), 2)).font = Font(bold=True, size=12, color="006100")

ws.cell(row=summary_row + 1, column=1, value="MIN").font = Font(bold=True)
ws.cell(row=summary_row + 1, column=11, value=round(min(cpu_spds), 2))
if gpu_spds:
    ws.cell(row=summary_row + 1, column=15, value=round(min(gpu_spds), 2))

ws.cell(row=summary_row + 2, column=1, value="MAX").font = Font(bold=True)
ws.cell(row=summary_row + 2, column=11, value=round(max(cpu_spds), 2))
if gpu_spds:
    ws.cell(row=summary_row + 2, column=15, value=round(max(gpu_spds), 2))

# Column widths
col_widths = [20, 8, 6, 16, 13, 13, 15, 15, 13, 10, 22, 15, 13, 10, 22, 35]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# --- Sheet 2: Sprint Review Highlights ---
ws2 = wb.create_sheet("Sprint Review")

ws2["A1"] = "HDBSCAN: Sprint Review Highlights"
ws2["A1"].font = Font(bold=True, size=16)
ws2["A2"] = ""

# Key metrics box
ws2["A3"] = "KEY METRICS"
ws2["A3"].font = Font(bold=True, size=12)
ws2["A4"] = f"CPU avg speedup: {np.mean(cpu_spds):.1f}x faster than scikit-learn"
ws2["A4"].font = Font(size=12)
ws2["A5"] = f"CPU best speedup: {max(cpu_spds):.1f}x (10k points, 2D)"
ws2["A5"].font = Font(size=12, color="006100")
if gpu_spds:
    ws2["A6"] = f"GPU avg speedup: {np.mean(gpu_spds):.1f}x faster than scikit-learn"
    ws2["A6"].font = Font(size=12)
    ws2["A7"] = f"GPU best speedup: {max(gpu_spds):.1f}x (5k points, 10D)"
    ws2["A7"].font = Font(size=12, color="006100")
ws2["A8"] = "Correctness: ARI >= 0.99 on all datasets (CPU: 13/13, GPU: 9/9 valid)"
ws2["A8"].font = Font(size=12)
ws2["A9"] = ""

# Best cases table
ws2["A10"] = "TOP 5 CPU SPEEDUPS"
ws2["A10"].font = Font(bold=True, size=12)

top_headers = ["Rank", "Dataset", "N points", "Dimensions", "sklearn (ms)", "oneDAL (ms)", "Speedup"]
for col, h in enumerate(top_headers, 1):
    cell = ws2.cell(row=11, column=col, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

sorted_cpu = sorted([r for r in results if r[10]], key=lambda x: x[10], reverse=True)
for i, r in enumerate(sorted_cpu[:5], 12):
    ws2.cell(row=i, column=1, value=i - 11)
    ws2.cell(row=i, column=2, value=r[0])
    ws2.cell(row=i, column=3, value=r[1])
    ws2.cell(row=i, column=4, value=r[2])
    ws2.cell(row=i, column=5, value=round(r[5], 0))
    ws2.cell(row=i, column=6, value=round(r[7], 0))
    cell = ws2.cell(row=i, column=7, value=f"{r[10]:.1f}x")
    cell.font = Font(bold=True, size=11)

ws2["A18"] = ""
ws2["A19"] = "TOP 5 GPU SPEEDUPS"
ws2["A19"].font = Font(bold=True, size=12)

for col, h in enumerate(top_headers, 1):
    cell = ws2.cell(row=20, column=col, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

sorted_gpu = sorted([r for r in results if r[14] and r[13] and r[13] >= 0.99],
                    key=lambda x: x[14], reverse=True)
for i, r in enumerate(sorted_gpu[:5], 21):
    ws2.cell(row=i, column=1, value=i - 20)
    ws2.cell(row=i, column=2, value=r[0])
    ws2.cell(row=i, column=3, value=r[1])
    ws2.cell(row=i, column=4, value=r[2])
    ws2.cell(row=i, column=5, value=round(r[5], 0))
    ws2.cell(row=i, column=6, value=round(r[11], 0))
    cell = ws2.cell(row=i, column=7, value=f"{r[14]:.1f}x")
    cell.font = Font(bold=True, size=11)

# Context section
ws2["A28"] = ""
ws2["A29"] = "CONTEXT"
ws2["A29"].font = Font(bold=True, size=12)
context_lines = [
    "- scikit-learn-intelex does NOT accelerate HDBSCAN (falls back to vanilla sklearn)",
    "- This implementation fills a key gap in oneDAL algorithm coverage",
    "- Algorithm: GEMM-based pairwise distances + Prim's MST + EOM cluster extraction",
    "- CPU: DAAL kernel with BlasInst (MKL) + threader_for + ISA dispatch (sse2/avx2/avx512)",
    "- GPU: SYCL kernels for distance matrix, host fallback for sequential MST",
    "- GPU limitation: >20k points with float32 can hit precision issues; >40k OOM on dist matrix",
    "- Tested across 13 datasets: 500 to 50,000 points, 2D to 10D",
]
for i, line in enumerate(context_lines, 30):
    ws2.cell(row=i, column=1, value=line)

for col in ["A", "B", "C", "D", "E", "F", "G"]:
    ws2.column_dimensions[col].width = 18
ws2.column_dimensions["A"].width = 75

# --- Sheet 3: Scaling Analysis ---
ws3 = wb.create_sheet("Scaling Analysis")
ws3["A1"] = "HDBSCAN Scaling: Speedup vs Dataset Size"
ws3["A1"].font = Font(bold=True, size=14)
ws3["A3"] = "CPU speedup improves with dataset size (BLAS GEMM amortization)"
ws3["A3"].font = Font(italic=True)

scale_headers = ["N (points)", "CPU Speedup", "GPU Speedup", "Observation"]
for col, h in enumerate(scale_headers, 1):
    ws3.cell(row=5, column=col, value=h).font = Font(bold=True)

scale_data = [
    (500, 2.18, 1.38, "Small: overhead dominated"),
    (1000, 2.63, 1.77, "Small: modest gains"),
    (2000, 5.37, 3.76, "Medium: GEMM kicks in"),
    (5000, 6.28, 5.74, "Medium: sweet spot for GPU"),
    (10000, 8.63, 3.47, "Large: CPU excels, GPU MST overhead"),
    (20000, 6.67, 4.18, "Large: strong CPU scaling"),
    (30000, 6.95, None, "Very large: GPU precision limit"),
    (50000, 7.56, None, "Very large: GPU OOM"),
]
for i, (n, cpu_s, gpu_s, obs) in enumerate(scale_data, 6):
    ws3.cell(row=i, column=1, value=n)
    ws3.cell(row=i, column=2, value=cpu_s)
    ws3.cell(row=i, column=3, value=gpu_s if gpu_s else "N/A")
    ws3.cell(row=i, column=4, value=obs)

for col in ["A", "B", "C", "D"]:
    ws3.column_dimensions[col].width = 20
ws3.column_dimensions["D"].width = 40

# Save
excel_path = os.path.join(OUTPUT_DIR, "hdbscan_benchmark_results.xlsx")
wb.save(excel_path)
print(f"Excel saved: {excel_path}")

# Also regenerate CSV with 50k included
csv_path = os.path.join(OUTPUT_DIR, "hdbscan_benchmark_results.csv")
with open(csv_path, "w") as f:
    f.write("Dataset,N,D,min_cluster_size,min_samples,sklearn_ms,sklearn_clusters,"
            "cpu_ms,cpu_clusters,cpu_ari,cpu_speedup,"
            "gpu_ms,gpu_clusters,gpu_ari,gpu_speedup,notes\n")
    for row in results:
        name, n, d, mcs, ms, sk_ms, sk_c, cpu_ms, cpu_c, cpu_ari, cpu_spd, gpu_ms, gpu_c, gpu_ari, gpu_spd, notes = row
        parts = [name, str(n), str(d), str(mcs), str(ms), f"{sk_ms:.1f}", str(sk_c)]
        parts.append(f"{cpu_ms:.1f}" if cpu_ms else "")
        parts.append(str(cpu_c) if cpu_c else "")
        parts.append(f"{cpu_ari:.4f}" if cpu_ari is not None else "")
        parts.append(f"{cpu_spd:.2f}" if cpu_spd else "")
        parts.append(f"{gpu_ms:.1f}" if gpu_ms else "")
        parts.append(str(gpu_c) if gpu_c else "")
        parts.append(f"{gpu_ari:.4f}" if gpu_ari is not None else "")
        parts.append(f"{gpu_spd:.2f}" if gpu_spd else "")
        parts.append(notes)
        f.write(",".join(parts) + "\n")
print(f"CSV saved: {csv_path}")
